# enriquecer_mecanismos_compra.py
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import re

BASE_DIR = Path("D:/Proyectos/P089 - Catastro Compras Cambio Climatico")
EXCEL_PATH = BASE_DIR / "Catastro_Cambio_Climatico_ChileCompra.xlsx"
CSV_PATH = BASE_DIR / "Catastro_Cambio_Climatico_ChileCompra.csv"

def extraer_mecanismo(row):
    cod = str(row['codigo_proceso']).upper().strip()
    tipo_reg = str(row['tipo_registro']).lower()
    
    # Extraer el tipo desde el sufijo (ej: 612227-1-LE26 o 2346-53-AG24 o 1057530-10-CM23)
    m = re.search(r'-([A-Z0-9]{2,3})\d*$', cod)
    suf = m.group(1)[:2] if m else ''
    
    if tipo_reg == 'licitacion':
        if suf in ['LP', 'LE', 'L1', 'LR', 'LQ', 'LS']:
            return 'Licitación Pública'
        elif suf in ['CO', 'B2']:
            return 'Licitación Privada'
        else:
            return 'Licitación Pública'
    else: # orden_compra
        if suf in ['AG']:
            return 'Compra Ágil (< 30 UTM)'
        elif suf in ['CM', 'CC', 'CD']:
            return 'Convenio Marco (Catálogo)'
        elif suf in ['TD', 'E2', 'SE']:
            return 'Trato Directo (Excepcional)'
        elif suf in ['LP', 'LE', 'L1', 'LR', 'LQ', 'LS']:
            return 'OC derivada de Licitación Pública'
        elif suf in ['CO', 'B2']:
            return 'OC derivada de Licitación Privada'
        else:
            return 'OC Ordinaria / Trato Directo'

def main():
    print("Cargando dataset para enriquecimiento de mecanismos de compra...")
    df = pd.read_excel(EXCEL_PATH, sheet_name="Catastro Completo")
    
    df['mecanismo_compra'] = df.apply(extraer_mecanismo, axis=1)
    
    print("\nDistribución por Mecanismo de Contratación:")
    print(df.groupby(['tipo_registro', 'mecanismo_compra']).size().to_string())
    
    # Guardar CSV enriquecido
    with open(CSV_PATH, 'w', encoding='utf-8-sig', newline='') as f:
        f.write('# dataset_id: P089_EVT_CHILECOMPRA_2007_2026\n')
        f.write('# data_architect: Eduardo Vega Toledo (evega.ap@gmail.com)\n')
        f.write('# project: P089 - Catastro Cambio Climatico Chile\n')
        df.to_csv(f, sep=';', index=False)
        
    print(f"\n[OK] CSV actualizado con mecanismo_compra: {CSV_PATH}")
    
    # Guardar Excel con nueva pestaña
    xl = pd.ExcelFile(EXCEL_PATH)
    sheets = {s: xl.parse(s) for s in xl.sheet_names}
    
    with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
        if 'Ficha Técnica y Autoría' in sheets:
            sheets['Ficha Técnica y Autoría'].to_excel(writer, sheet_name='Ficha Técnica y Autoría', index=False)
            
        df.to_excel(writer, sheet_name='Catastro Completo', index=False)
        
        # Pestaña de Mecanismos de Compra
        resumen_mec = df.groupby(['tipo_registro', 'mecanismo_compra']).agg(
            total_procesos=('codigo_proceso', 'count'),
            monto_millones=('monto_pesos', lambda x: round(pd.to_numeric(x, errors='coerce').fillna(0).sum() / 1e6, 2))
        ).reset_index()
        resumen_mec.to_excel(writer, sheet_name='Resumen Mecanismos Compra', index=False)
        
        for s_name in ['Resumen Nivel Institucional', 'Resumen Subcategorias', 'Detalle GOREs', 'Top 50 Organismos', 'Terminos Gatillantes']:
            if s_name in sheets:
                sheets[s_name].to_excel(writer, sheet_name=s_name, index=False)
                
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Ficha Técnica y Autoría']
    header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    for col_num in range(1, 3):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 110
    wb.save(EXCEL_PATH)
    
    print(f"[OK] Excel actualizado con nueva pestaña 'Resumen Mecanismos Compra': {EXCEL_PATH}")

if __name__ == "__main__":
    main()
